"""Build a locked benchmark bundle from an analyst-owned Excel reference.

This module is a build-time adapter.  The generated truth JSON is consumed by
``LockedBenchmarkScorer`` only after blind analysis output has been archived.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any, Iterable

from .contracts import MANIFEST_SCHEMA_VERSION, TRUTH_SCHEMA_VERSION, sha256_file


INSULIN_REQUIRED_SHEETS = (
    "Anchor_Reference",
    "Kinase_Reference",
    "Temporal_Layers",
    "Ambiguous_Sites",
    "Scoring_Template",
    "Benchmark_Rules",
)

INSULIN_OPTIONAL_V2_SHEETS = (
    "Protein_Effectors",
    "Cross_Layer_Relations",
    "Mechanism_Chains",
    "Counterexamples",
)

_HEADER_TOKENS = {
    "Anchor_Reference": "Anchor_ID",
    "Kinase_Reference": "Kinase_or_complex",
    "Temporal_Layers": "Window_ID",
    "Ambiguous_Sites": "Site_or_pattern",
    "Scoring_Template": "Anchor_ID",
    "Benchmark_Rules": "Rule_ID",
    "Protein_Effectors": "Effector_ID",
    "Cross_Layer_Relations": "Relation_ID",
    "Mechanism_Chains": "Chain_ID",
    "Counterexamples": "Counterexample_ID",
}

# The analyst-owned Anchor Reference permits an empty Rat_site cell for a
# human-reference-only anchor.  Excel can serialise that internal empty cell by
# omitting it from a ragged row, which would otherwise shift every later field
# one column to the left.  Keep this schema rule declarative and sheet-scoped;
# no row IDs or biological facts are hard-coded here.
_OPTIONAL_RAGGED_HEADERS = {
    "Anchor_Reference": ("Rat_site",),
}


def build_insulin_locked_reference(
    workbook_path: str | Path,
    output_dir: str | Path,
    *,
    dataset_id: str = "insulin_signaling_v1",
) -> tuple[Path, Path]:
    """Convert the provided insulin workbook to a versioned truth/manifest pair."""

    from openpyxl import load_workbook

    source = Path(workbook_path).resolve()
    destination = Path(output_dir).resolve()
    locked_dir = destination / "locked_truth"
    locked_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source, data_only=True, read_only=True)
    missing = [sheet for sheet in INSULIN_REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError("insulin workbook is missing sheets: " + ", ".join(missing))

    sheets = {
        sheet: _worksheet_records(
            workbook[sheet].iter_rows(values_only=True),
            header_token=_HEADER_TOKENS[sheet],
            optional_ragged_headers=_OPTIONAL_RAGGED_HEADERS.get(sheet, ()),
        )
        for sheet in INSULIN_REQUIRED_SHEETS
    }
    for sheet in INSULIN_OPTIONAL_V2_SHEETS:
        if sheet not in workbook.sheetnames:
            continue
        sheets[sheet] = _worksheet_records(
            workbook[sheet].iter_rows(values_only=True),
            header_token=_HEADER_TOKENS[sheet],
            optional_ragged_headers=_OPTIONAL_RAGGED_HEADERS.get(sheet, ()),
        )
    anchors = sheets["Anchor_Reference"]
    truth = {
        "schema_version": TRUTH_SCHEMA_VERSION,
        "dataset_id": dataset_id,
        "source_workbook_name": source.name,
        "source_workbook_sha256": sha256_file(source),
        "anchors": anchors,
        "kinase_reference": sheets["Kinase_Reference"],
        "temporal_layers": sheets["Temporal_Layers"],
        "ambiguous_sites": sheets["Ambiguous_Sites"],
        "scoring_template": sheets["Scoring_Template"],
        "benchmark_rules": sheets["Benchmark_Rules"],
        "additive_v2_reference": {
            "protein_effectors": sheets.get("Protein_Effectors", []),
            "cross_layer_relations": sheets.get("Cross_Layer_Relations", []),
            "mechanism_chains": sheets.get("Mechanism_Chains", []),
            "counterexamples": sheets.get("Counterexamples", []),
            "source_sheets_present": [
                sheet for sheet in INSULIN_OPTIONAL_V2_SHEETS if sheet in workbook.sheetnames
            ],
        },
    }
    truth_path = locked_dir / f"{dataset_id}.truth.json"
    _write_json(truth_path, truth)
    manifest = {
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "dataset_id": dataset_id,
        "display_name": "Insulin signaling phospho-kinase benchmark v1",
        "locked_truth_bundle": str(Path("locked_truth") / truth_path.name),
        "locked_truth_sha256": sha256_file(truth_path),
        "source_reference": {
            "kind": "analyst_provided_workbook",
            "workbook_sha256": truth["source_workbook_sha256"],
        },
        "production_contract": {
            "id": "tmm_full_temporal.v1",
            "layers": ["product_pipeline.v0", "temporal_science.v1"],
            "temporal_contract": "dynamics_v1",
            "kinase_scoring_method": "tmm_full_temporal",
            "tmm_guard_policy": "group_share",
            "representation_learning_in_primary_score": False,
        },
        "blind_policy": {
            "stimulus_hidden_from_analysis_runtime": True,
            "research_question_hidden_from_analysis_runtime": True,
            "truth_available_to_scorer_only": True,
            "rag_policy": "disabled_for_strict_primary",
            "cell_context_policy": "lineage_only",
        },
        "score_config": {
            "evidence_tier_weights": {"Tier 1": 2, "Tier 2": 1},
            "component_weights": {
                "detectable_anchor_recall": 0.25,
                "regulated_anchor_recall": 0.25,
                "direction_accuracy": 0.20,
                "peak_window_accuracy": 0.20,
                "chain_completeness": 0.10,
            },
            "novel_tier_policy": "Tier 3/4 and de_novo are discovery-only; never score canonical accuracy",
            "site_mapping_requirement": "sequence_isoform_species",
        },
    }
    manifest_path = destination / f"{dataset_id}.manifest.json"
    _write_json(manifest_path, manifest)
    return manifest_path, truth_path


def _worksheet_records(
    rows: Iterable[tuple[Any, ...]],
    *,
    header_token: str,
    optional_ragged_headers: tuple[str, ...] = (),
) -> list[dict[str, Any]]:
    materialized = list(rows)
    header_index = next(
        (
            index
            for index, row in enumerate(materialized)
            if header_token in {str(value or "").strip() for value in row}
        ),
        None,
    )
    if header_index is None:
        raise ValueError(f"could not locate {header_token!r} header in workbook sheet")
    headers = materialized[header_index]
    if not any(headers):
        return []
    normalized_headers = [str(value or "").strip() for value in headers]
    records: list[dict[str, Any]] = []
    for row in materialized[header_index + 1 :]:
        if not any(value not in (None, "") for value in row):
            continue
        normalized_row = _restore_optional_ragged_cells(
            row,
            normalized_headers,
            optional_ragged_headers,
        )
        record = {
            header: _json_value(value)
            for header, value in zip(normalized_headers, normalized_row)
            if header
        }
        records.append(record)
    return records


def _restore_optional_ragged_cells(
    row: tuple[Any, ...],
    headers: list[str],
    optional_ragged_headers: tuple[str, ...],
) -> tuple[Any, ...]:
    """Restore omitted internal optional cells without inventing data.

    A row may be shorter than its header only because a declared optional
    header was omitted.  Insert ``None`` at the matching schema position.  If
    the deficit cannot be fully explained, leave the row untouched rather than
    guessing where a biological value belongs.
    """

    values = tuple(row)
    if not optional_ragged_headers:
        return values

    optional_positions = [
        headers.index(header)
        for header in optional_ragged_headers
        if header in headers
    ]
    if not optional_positions:
        return values

    if _row_alignment_is_valid(values, headers):
        return values

    restored = list(values)
    for position in sorted(optional_positions):
        restored.insert(position, None)
    while len(restored) > len(headers) and restored[-1] in (None, ""):
        restored.pop()
    candidate = tuple(restored)
    return candidate if _row_alignment_is_valid(candidate, headers) else values


def _row_alignment_is_valid(values: tuple[Any, ...], headers: list[str]) -> bool:
    """Validate only declared structural signals, never biological values."""

    if len(values) != len(headers):
        return False
    if "Evidence_tier" not in headers:
        return True
    value = values[headers.index("Evidence_tier")]
    return str(value or "").strip().startswith("Tier ")


def _json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
