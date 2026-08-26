from benchmarking.v2_scorer import score_additive_v2
import hashlib
import json

from benchmarking.v2_scorer import score_additive_v2
from benchmarking.v2_truth_adapter import build_additive_v2_truth


def test_v1_truth_adapter_does_not_invent_cross_layer_reference() -> None:
    truth = build_additive_v2_truth(
        {
            "dataset_id": "x",
            "source_workbook_sha256": "abc",
            "kinase_reference": [
                {
                    "Kinase_ID": "K1",
                    "Kinase_or_complex": "K1",
                    "Expected_activity_direction": "Up",
                    "Expected_time": "Early",
                    "Direct_or_preferred_outputs": "P1; P2",
                }
            ],
            "temporal_layers": [],
            "ambiguous_sites": [],
        }
    )
    assert truth["cross_layer_reference"] == []
    assert truth["evaluability"]["cross_layer"] == "not_evaluable_missing_optional_sheet"
    assert truth["mechanism_reference"][0]["required_output_tokens"] == ["P1", "P2"]


def test_additive_scorer_uses_not_evaluable_for_zero_data_anchor() -> None:
    truth = build_additive_v2_truth(
        {
            "dataset_id": "x",
            "kinase_reference": [{"Kinase_ID": "K1", "Kinase_or_complex": "K1", "Expected_time": "Early"}],
            "temporal_layers": [],
            "ambiguous_sites": [],
        }
    )
    score = score_additive_v2(
        {
            "v2_extensions": {
                "kinase_timing_predictions": [
                    {"kinase": "K1", "data_anchored": False, "peak_timepoint": "5min"}
                ],
                "cross_layer_edges": [],
                "mechanism_chains": [],
                "mechanism_counterevidence": [],
            }
        },
        truth,
    )
    assert score["kinase_evidence_v2"]["timing_status"] == "not_evaluable"
    assert score["kinase_evidence_v2"]["metrics"]["timing_accuracy_data_anchored"] is None
    assert score["score_isolation"]["combined_weighted_score"] is None


def test_optional_truth_normalization_and_independent_recovery_metrics() -> None:
    parent = {
        "dataset_id": "x",
        "kinase_reference": [],
        "temporal_layers": [],
        "ambiguous_sites": [],
        "additive_v2_reference": {
            "protein_effectors": [
                {"Effector_ID": "E1", "Gene": "gene1", "Expected_peak": "5min", "Expected_direction": "up"}
            ],
            "cross_layer_relations": [
                {
                    "Relation_ID": "R1",
                    "Source_wave_ID": "TW-01",
                    "Target_gene": "gene1",
                    "Expected_direction": "source_precedes_target",
                    "Minimum_peak_lag_minutes": 1,
                    "Maximum_peak_lag_minutes": 10,
                }
            ],
            "mechanism_chains": [
                {"Chain_ID": "C1", "Kinase_or_complex": "K1", "Target_gene": "gene1"}
            ],
            "counterexamples": [
                {"Counterexample_ID": "X1", "Chain_ID": "K1__TW-01__GENE1", "Expected_status": "insufficient_evidence"}
            ],
        },
    }
    truth = build_additive_v2_truth(parent)
    expected_parent_hash = hashlib.sha256(
        json.dumps(parent, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    assert truth["parent_v1_truth_sha256"] == expected_parent_hash
    assert truth["mechanism_reference"][0]["required_output_tokens"] == ["GENE1"]
    score = score_additive_v2(
        {
            "v2_extensions": {
                "kinase_timing_predictions": [],
                "protein_time_series": [
                    {"gene": "GENE1", "peak_timepoint": "5min", "peak_direction": "up"}
                ],
                "cross_layer_edges": [
                    {
                        "edge_id": "TW-01__GENE1",
                        "source_wave_id": "TW-01",
                        "target_gene": "GENE1",
                        "direction": "source_precedes_target",
                        "peak_lag_minutes": 5,
                        "eligible_for_mechanism_chain": True,
                    }
                ],
                "mechanism_chains": [
                    {
                        "chain_id": "K1__TW-01__GENE1",
                        "kinase": "K1",
                        "target_gene": "GENE1",
                        "mechanism_status": "evidence_supported_mechanism_candidate",
                    }
                ],
                "mechanism_counterevidence": [
                    {"chain_id": "K1__TW-01__GENE1", "status": "insufficient_evidence"}
                ],
            }
        },
        truth,
    )
    assert score["protein_effectors_v2"]["reference_recovery"] == 1.0
    assert score["cross_layer_v2"]["reference_recovery"] == 1.0
    assert score["mechanism_v2"]["reference_recovery"] == 1.0
    assert score["refutation_v2"]["refutation_sensitivity"] == 1.0


def test_blank_template_contains_headers_but_no_truth_rows(tmp_path) -> None:
    from openpyxl import load_workbook
    from benchmarking.v2_truth_template import SHEET_HEADERS, create_blank_additive_v2_template

    path = create_blank_additive_v2_template(tmp_path / "additive-v2-template.xlsx")
    workbook = load_workbook(path, data_only=True, read_only=True)
    for sheet_name, headers in SHEET_HEADERS.items():
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        assert list(rows[0]) == headers
        assert len(rows) == 1
