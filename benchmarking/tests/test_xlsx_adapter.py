from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from benchmarking.contracts import BenchmarkManifest, load_locked_truth_bundle
from benchmarking.xlsx_adapter import INSULIN_REQUIRED_SHEETS, build_insulin_locked_reference


def _workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    rows = {
        "Anchor_Reference": [["Anchor_ID", "Evidence_tier", "Branch", "Expected_p_direction", "Expected_peak_window", "Benchmark_truth_use"], ["A001", "Tier 1", "INSR", "Up", "1–15 min", "Positive truth"]],
        "Kinase_Reference": [["Kinase_ID", "Kinase_or_complex"], ["K01", "AKT1"]],
        "Temporal_Layers": [["Window_ID", "Temporal_layer"], ["W01", "Immediate"]],
        "Ambiguous_Sites": [["Site_or_pattern", "Benchmark_policy"], ["A001", "none"]],
        "Scoring_Template": [["Anchor_ID", "Weight"], ["A001", 2]],
        "Benchmark_Rules": [["Rule_ID", "Rule"], ["R01", "Tier 1/2"]],
    }
    for name in INSULIN_REQUIRED_SHEETS:
        sheet = workbook.create_sheet(name)
        sheet.append([f"{name} title row"])
        for row in rows[name]:
            sheet.append(row)
    workbook.save(path)


def test_adapter_builds_manifest_and_locked_truth(tmp_path: Path) -> None:
    source = tmp_path / "insulin.xlsx"
    _workbook(source)
    manifest_path, truth_path = build_insulin_locked_reference(source, tmp_path / "dataset")
    manifest = BenchmarkManifest.load(manifest_path)
    truth = load_locked_truth_bundle(manifest)
    assert truth_path.is_file()
    assert truth["anchors"][0]["Anchor_ID"] == "A001"
    assert manifest.production_contract["id"] == "tmm_full_temporal.v1"
    assert json.loads(manifest_path.read_text())["blind_policy"]["truth_available_to_scorer_only"] is True


def test_adapter_restores_declared_optional_anchor_cell_without_shifting_fields(tmp_path: Path) -> None:
    source = tmp_path / "insulin-ragged.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    rows = {
        "Anchor_Reference": [
            ["Anchor_ID", "Gene", "Human_site", "Rat_site", "Mapping_status", "Evidence_tier", "Benchmark_truth_use"],
            ["A001", "INSR", "Y960", "IRS docking motif", "Tier 2", "Positive truth if measurable"],
        ],
        "Kinase_Reference": [["Kinase_or_complex"], ["INSR"]],
        "Temporal_Layers": [["Window_ID"], ["W01"]],
        "Ambiguous_Sites": [["Site_or_pattern"], ["Y960"]],
        "Scoring_Template": [["Anchor_ID"], ["A001"]],
        "Benchmark_Rules": [["Rule_ID"], ["R01"]],
    }
    for name in INSULIN_REQUIRED_SHEETS:
        sheet = workbook.create_sheet(name)
        sheet.append([f"{name} title row"])
        for row in rows[name]:
            sheet.append(row)
    # Reproduce the source workbook's ragged-row encoding: an empty internal
    # Rat_site cell is represented as a shifted value plus a trailing blank.
    anchor_sheet = workbook["Anchor_Reference"]
    anchor_sheet.cell(row=3, column=7).value = None
    workbook.save(source)

    manifest_path, _ = build_insulin_locked_reference(source, tmp_path / "dataset")
    truth = load_locked_truth_bundle(BenchmarkManifest.load(manifest_path))
    anchor = truth["anchors"][0]
    assert anchor["Rat_site"] is None
    assert anchor["Mapping_status"] == "IRS docking motif"
    assert anchor["Evidence_tier"] == "Tier 2"
    assert anchor["Benchmark_truth_use"] == "Positive truth if measurable"
