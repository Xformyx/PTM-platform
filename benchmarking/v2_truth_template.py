"""Blank analyst-owned additive-v2 workbook extension schema."""

from __future__ import annotations

from pathlib import Path


SHEET_HEADERS = {
    "Protein_Effectors": [
        "Effector_ID",
        "Gene",
        "Expected_peak",
        "Expected_direction",
        "Evidence_tier",
        "Reference",
        "Notes",
    ],
    "Cross_Layer_Relations": [
        "Relation_ID",
        "Source_wave_ID",
        "Target_gene",
        "Expected_direction",
        "Minimum_peak_lag_minutes",
        "Maximum_peak_lag_minutes",
        "Evidence_tier",
        "Reference",
        "Notes",
    ],
    "Mechanism_Chains": [
        "Chain_ID",
        "Kinase_ID",
        "Kinase_or_complex",
        "Wave_ID",
        "Target_gene",
        "Required_output_tokens",
        "Expected_direction",
        "Expected_time",
        "Evidence_tier",
        "Reference",
        "Notes",
    ],
    "Counterexamples": [
        "Counterexample_ID",
        "Chain_ID",
        "Kinase_or_complex",
        "Target_gene",
        "Expected_status",
        "Exclusion_reason",
        "Reference",
        "Notes",
    ],
}


def create_blank_additive_v2_template(output: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    target = Path(output)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["Benchmark additive-v2 runner-only truth extension"])
    readme.append(["This workbook is intentionally blank. An independent analyst must author every biological reference row."])
    readme.append(["Never populate these sheets from algorithm predictions, raw-data candidates, RAG, LLM output, or the stimulus-aware report."])
    readme.append(["Merge completed sheets into a copy of the locked v1 workbook before running xlsx_adapter."])
    for sheet_name, headers in SHEET_HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[chr(64 + index)].width = max(16, min(40, len(header) + 4))
    workbook.save(target)
    return target
