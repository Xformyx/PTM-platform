"""Derive runner-only optional benchmark truth and a separate extension workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from benchmarking.workbook_optional_truth_derivation import (
    derive_optional_truth_from_workbook,
    provenance_summary,
)
from benchmarking.v2_truth_adapter import build_additive_v2_truth


OPTIONAL_HEADERS = {
    "Protein_Effectors": ["Effector_ID", "Gene", "Expected_peak", "Expected_direction", "Evidence_tier", "Reference", "Notes"],
    "Cross_Layer_Relations": ["Relation_ID", "Source_wave_ID", "Target_gene", "Expected_direction", "Minimum_peak_lag_minutes", "Maximum_peak_lag_minutes", "Evidence_tier", "Reference", "Notes"],
    "Mechanism_Chains": ["Chain_ID", "Kinase_ID", "Kinase_or_complex", "Target_gene", "Required_output_tokens", "Expected_direction", "Expected_time", "Evidence_tier", "Reference", "Notes", "Reference_origin"],
    "Counterexamples": ["Counterexample_ID", "Chain_ID", "Kinase_or_complex", "Target_gene", "Expected_status", "Exclusion_reason", "Reference", "Notes"],
}


def _write_extension_workbook(source: Path, output: Path, derived_truth: dict) -> None:
    workbook = load_workbook(source)
    for sheet_name, headers in OPTIONAL_HEADERS.items():
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        if sheet_name == "Mechanism_Chains":
            for row in derived_truth["additive_v2_reference"]["mechanism_chains"]:
                sheet.append([row.get(header) for header in headers])
    provenance_name = "Optional_Truth_Provenance"
    if provenance_name in workbook.sheetnames:
        del workbook[provenance_name]
    provenance = workbook.create_sheet(provenance_name)
    provenance.append(["Field", "Value"])
    for key, value in derived_truth["workbook_optional_truth_derivation"].items():
        provenance.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value])
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--base-truth", required=True)
    parser.add_argument("--output-truth", required=True)
    parser.add_argument("--output-workbook", required=True)
    parser.add_argument("--summary-output", required=True)
    args = parser.parse_args()
    base_truth = json.loads(Path(args.base_truth).read_text(encoding="utf-8"))
    derived = derive_optional_truth_from_workbook(base_truth, workbook_path=args.workbook)
    additive_truth = build_additive_v2_truth(derived)
    output_truth = Path(args.output_truth)
    output_truth.parent.mkdir(parents=True, exist_ok=True)
    output_truth.write_text(json.dumps(additive_truth, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    _write_extension_workbook(Path(args.workbook), Path(args.output_workbook), derived)
    summary = {
        **provenance_summary(derived),
        "additive_truth_sha256": additive_truth["truth_sha256"],
        "mechanism_evaluability": additive_truth["evaluability"]["mechanism"],
    }
    Path(args.summary_output).write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
